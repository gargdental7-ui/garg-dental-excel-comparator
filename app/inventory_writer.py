"""Writes the Inventory Movement Analyzer output workbook."""
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from .inventory_analyzer import NO_MOVEMENT, SLOW_MOVING, FAST_MOVING
from .report_style import autofit_columns, style_header_row

PRODUCT_HEADER = [
    "Code",
    "Description",
    "Unit",
    "Opening",
    "Received",
    "Delivered",
    "Balance",
    "Movement Ratio",
    "Movement Classification",
    "Inventory Value",
]


def _product_row(p, include_value):
    row = [
        p.code,
        p.description,
        p.unit,
        p.opening,
        p.received,
        p.delivered,
        p.balance,
        round(p.movement_ratio, 4),
        p.classification,
        round(p.inventory_value, 2) if p.inventory_value is not None else None,
    ]
    return row if include_value else row[:-1]


def _write_products_sheet(ws, products, include_value):
    header = PRODUCT_HEADER if include_value else PRODUCT_HEADER[:-1]
    ws.append(header)
    for p in products:
        ws.append(_product_row(p, include_value))
    style_header_row(ws, len(header))
    autofit_columns(ws, len(header))


def _write_exceptions_sheet(ws, exceptions):
    header = ["Code", "Reason", "Excel Row"]
    ws.append(header)
    for exc in exceptions:
        ws.append([exc.code, exc.reason, exc.excel_row_index])
    style_header_row(ws, len(header))
    autofit_columns(ws, len(header))


def write_inventory_report(output_path, result):
    """Write INVENTORY_ANALYSIS, NO_MOVEMENT, SLOW_MOVING, FAST_MOVING,
    STOCK_EXCEPTIONS (+ HIGH_VALUE_RISK if value data is available)."""
    wb = Workbook()
    include_value = result.has_value_data

    ws_all = wb.active
    ws_all.title = "INVENTORY_ANALYSIS"
    _write_products_sheet(ws_all, result.products, include_value)

    ws_no_move = wb.create_sheet("NO_MOVEMENT")
    _write_products_sheet(
        ws_no_move, [p for p in result.products if p.classification == NO_MOVEMENT], include_value
    )

    ws_slow = wb.create_sheet("SLOW_MOVING")
    _write_products_sheet(
        ws_slow, [p for p in result.products if p.classification == SLOW_MOVING], include_value
    )

    ws_fast = wb.create_sheet("FAST_MOVING")
    _write_products_sheet(
        ws_fast, [p for p in result.products if p.classification == FAST_MOVING], include_value
    )

    ws_exceptions = wb.create_sheet("STOCK_EXCEPTIONS")
    _write_exceptions_sheet(ws_exceptions, result.exceptions)

    if include_value:
        ws_risk = wb.create_sheet("HIGH_VALUE_RISK")
        risky = [
            p
            for p in result.products
            if p.classification in (NO_MOVEMENT, SLOW_MOVING) and (p.inventory_value or 0) > 0
        ]
        risky.sort(key=lambda p: p.inventory_value or 0, reverse=True)
        _write_products_sheet(ws_risk, risky, include_value)

    output_path = Path(output_path)
    wb.save(str(output_path))
    return output_path


def default_output_filename():
    date_str = datetime.now().strftime("%Y-%m-%d")
    return f"Garg_Dental_Inventory_Analysis_{date_str}.xlsx"
