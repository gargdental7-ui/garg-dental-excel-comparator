"""Shared openpyxl formatting helpers for the Collection and Inventory
report writers (bold header, frozen header row, autofilter, column widths)."""
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def style_header_row(ws, ncols, row=1):
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).font = Font(bold=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate
    ws.auto_filter.ref = f"A{row}:{get_column_letter(ncols)}{row}"


def autofit_columns(ws, ncols, min_width=10, max_width=40):
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        best = min_width
        for cell in ws[letter]:
            if cell.value is not None:
                best = max(best, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[letter].width = best
