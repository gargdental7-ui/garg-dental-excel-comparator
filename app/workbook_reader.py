"""Reads an Excel workbook into a lightweight, comparison-ready structure."""
import logging
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from .exceptions import WorkbookOpenError
from .header_detector import detect_header
from .normalizers import is_no_group_row, normalize_code

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm")


@dataclass
class RowRecord:
    excel_row_index: int
    code_raw: object
    code_key: str
    values: list


class SheetData:
    def __init__(self, path, workbook, worksheet, header_info, rows):
        self.path = Path(path)
        self.workbook = workbook
        self.worksheet = worksheet
        self.header_row_start = header_info.header_row_start
        self.header_row_count = header_info.header_row_count
        self.headers = header_info.headers
        self.num_columns = header_info.num_columns
        self.rows = rows


def _is_blank_row(values):
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in values)


def load_sheet(path, file_label):
    """Load the first worksheet of an Excel file and return a SheetData ready
    for comparison. Raises AppError subclasses with human-readable messages."""
    path = Path(path)
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise WorkbookOpenError(str(path))

    try:
        workbook = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as exc:
        logger.exception("Failed to open workbook: %s", path)
        raise WorkbookOpenError(str(path)) from exc

    worksheet = workbook.worksheets[0]
    header_info = detect_header(worksheet, file_label)
    code_idx = next(i for i, h in enumerate(header_info.headers) if h.lower() == "code")

    data_start = header_info.header_row_start + header_info.header_row_count
    rows = []
    for row_idx in range(data_start, worksheet.max_row + 1):
        raw_values = [
            worksheet.cell(row=row_idx, column=c).value
            for c in range(1, header_info.num_columns + 1)
        ]
        if _is_blank_row(raw_values):
            continue
        code_raw = raw_values[code_idx] if code_idx < len(raw_values) else None
        code_key = normalize_code(code_raw)
        if code_key is None or is_no_group_row(code_key):
            continue
        rows.append(RowRecord(row_idx, code_raw, code_key, raw_values))

    return SheetData(path, workbook, worksheet, header_info, rows)
