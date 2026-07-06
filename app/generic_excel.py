"""Generic Excel import used by the Collection and Inventory analyzers.

Unlike the Comparator's workbook_reader (which requires a literal "Code"
header), these tools accept arbitrary business reports with unknown column
names, so headers are detected heuristically and the user maps columns to
logical fields afterwards.
"""
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import openpyxl

from .exceptions import EmptyWorkbookError, GenericHeaderDetectionError, WorkbookOpenError

SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm")
MAX_HEADER_SCAN_ROWS = 50


@dataclass
class GenericSheetData:
    sheet_name: str
    headers: list
    header_row_start: int
    header_row_count: int
    num_columns: int
    rows: list
    row_excel_indexes: list


def open_workbook(path, file_label):
    """Open an uploaded workbook for reading. Never modifies the source file."""
    path = Path(path)
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise WorkbookOpenError(str(path))
    try:
        workbook = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as exc:
        raise WorkbookOpenError(str(path)) from exc
    if not workbook.worksheets:
        raise EmptyWorkbookError(file_label)
    return workbook


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


def _row_raw_values(ws, row_idx, max_col):
    return [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]


def _looks_numeric(raw_value):
    if isinstance(raw_value, bool):
        return False
    if isinstance(raw_value, (int, float, date, datetime)):
        return True
    if raw_value is None:
        return False
    text = str(raw_value).strip()
    if text == "":
        return False
    cleaned = re.sub(r"[,₹$%()]", "", text).strip()
    cleaned = re.sub(r"(?i)^rs\.?\s*", "", cleaned).strip()
    if cleaned in ("", "-"):
        return False
    try:
        float(cleaned)
        return True
    except ValueError:
        return False


def _dedupe_headers(headers):
    seen = {}
    result = []
    for i, h in enumerate(headers, start=1):
        name = h if h else f"Column {i}"
        base = name
        counter = 2
        while name in seen:
            name = f"{base} ({counter})"
            counter += 1
        seen[name] = True
        result.append(name)
    return result


def detect_generic_header(ws, file_label, max_scan_rows=MAX_HEADER_SCAN_ROWS):
    """Locate the header row by scanning for a text-heavy row immediately
    followed by data (numbers/dates), or - for the known Garg Dental
    two-row layout - a text-heavy row followed by a mostly-blank sub-header
    row (e.g. "Qty") and then data."""
    if ws.max_row == 0 or ws.max_column == 0:
        raise GenericHeaderDetectionError(file_label)

    max_col = ws.max_column
    scan_limit = min(ws.max_row, max_scan_rows)

    for row_idx in range(1, scan_limit + 1):
        raw = _row_raw_values(ws, row_idx, max_col)
        cleaned = [_clean(v) for v in raw]
        non_blank = [v for v in cleaned if v]
        if len(non_blank) < 2:
            continue

        numeric_count = sum(1 for v in raw if _looks_numeric(v))
        text_ratio = 1 - (numeric_count / len(non_blank))
        if text_ratio < 0.5:
            continue

        # Two-row Garg-style sub-headers must sit directly under the header
        # row (no blank spacer), so check adjacency before anything else.
        if row_idx + 1 <= ws.max_row:
            next_raw = _row_raw_values(ws, row_idx + 1, max_col)
            next_cleaned = [_clean(v) for v in next_raw]
            next_non_blank = [v for v in next_cleaned if v]
            if next_non_blank:
                looks_like_subheader = (not next_cleaned[0]) and len(next_non_blank) < len(non_blank)
                if looks_like_subheader and row_idx + 2 <= ws.max_row:
                    after_raw = _row_raw_values(ws, row_idx + 2, max_col)
                    after_non_blank = [v for v in after_raw if _clean(v)]
                    if after_non_blank:
                        return row_idx, cleaned, max_col, 2

        # Otherwise look ahead (tolerating blank spacer rows) for the first
        # row that looks like data, confirming this candidate is a real
        # header rather than another title/label line.
        lookahead_limit = min(ws.max_row, row_idx + 5)
        for probe in range(row_idx + 1, lookahead_limit + 1):
            probe_raw = _row_raw_values(ws, probe, max_col)
            probe_cleaned = [_clean(v) for v in probe_raw]
            probe_non_blank = [v for v in probe_cleaned if v]
            if not probe_non_blank:
                continue
            probe_numeric_count = sum(1 for v in probe_raw if _looks_numeric(v))
            if (probe_numeric_count / len(probe_non_blank)) > 0.3:
                return row_idx, cleaned, max_col, 1
            break

    raise GenericHeaderDetectionError(file_label)


def _is_blank_row(values):
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in values)


def load_generic_sheet(workbook, sheet_name, file_label, max_scan_rows=MAX_HEADER_SCAN_ROWS):
    worksheet = workbook[sheet_name]
    row_idx, header_values, max_col, header_row_count = detect_generic_header(
        worksheet, file_label, max_scan_rows
    )

    last_named = 0
    for i, h in enumerate(header_values, start=1):
        if h:
            last_named = i
    max_col = last_named
    headers = _dedupe_headers(header_values[:max_col])

    data_start = row_idx + header_row_count
    rows = []
    row_excel_indexes = []
    for r in range(data_start, worksheet.max_row + 1):
        raw_values = [worksheet.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if _is_blank_row(raw_values):
            continue
        rows.append(dict(zip(headers, raw_values)))
        row_excel_indexes.append(r)

    return GenericSheetData(
        sheet_name=sheet_name,
        headers=headers,
        header_row_start=row_idx,
        header_row_count=header_row_count,
        num_columns=max_col,
        rows=rows,
        row_excel_indexes=row_excel_indexes,
    )
