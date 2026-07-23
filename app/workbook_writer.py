"""Writes the comparison result workbook, copying rows and styling directly
from the source worksheets so complete OMS rows are preserved byte-for-byte
rather than reconstructed from selected fields."""
import logging
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .report_style import autofit_columns, style_header_row

logger = logging.getLogger(__name__)

# Column Comparison Report Mode conditional-formatting fills. Light shades
# throughout so black text stays readable - "Blue"/"Red" in the spec are
# interpreted as light blue/light red row fills for that reason.
_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_LIGHT_GREEN = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
_LIGHT_BLUE = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_LIGHT_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_ORANGE = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")


def _copy_row(dest_ws, dest_row, src_ws, src_row, num_cols):
    for col in range(1, num_cols + 1):
        src_cell = src_ws.cell(row=src_row, column=col)
        dest_cell = dest_ws.cell(row=dest_row, column=col, value=src_cell.value)
        if src_cell.has_style:
            # openpyxl wraps font/fill/border/alignment in a StyleProxy that
            # is not hashable; copy() unwraps it to the real style object
            # before assigning to another cell.
            dest_cell.font = copy(src_cell.font)
            dest_cell.border = copy(src_cell.border)
            dest_cell.fill = copy(src_cell.fill)
            dest_cell.alignment = copy(src_cell.alignment)
            dest_cell.number_format = src_cell.number_format


def _copy_presentation(dest_ws, src_ws, num_cols):
    for col in range(1, num_cols + 1):
        letter = get_column_letter(col)
        src_dim = src_ws.column_dimensions.get(letter)
        if src_dim and src_dim.width:
            dest_ws.column_dimensions[letter].width = src_dim.width
    if src_ws.freeze_panes:
        dest_ws.freeze_panes = src_ws.freeze_panes


def _write_sheet(dest_ws, source_sheet, row_records):
    num_cols = source_sheet.num_columns
    for offset in range(source_sheet.header_row_count):
        dest_r = offset + 1
        src_r = source_sheet.header_row_start + offset
        _copy_row(dest_ws, dest_r, source_sheet.worksheet, src_r, num_cols)

    dest_row = source_sheet.header_row_count + 1
    for row in row_records:
        _copy_row(dest_ws, dest_row, source_sheet.worksheet, row.excel_row_index, num_cols)
        dest_row += 1

    _copy_presentation(dest_ws, source_sheet.worksheet, num_cols)


def _write_field_changes_sheet(dest_ws, cell_differences):
    dest_ws.append(["Code", "Column", "Old Value (Current File)", "New Value (OMS File)"])
    for cell in dest_ws[1]:
        cell.font = Font(bold=True)
    for diff in cell_differences:
        dest_ws.append([diff.code, diff.column, diff.old_value, diff.new_value])


def write_result(output_path, current_sheet, oms_sheet, result):
    """Write FIELD_CHANGES + DIFFERENCES (+ NEW_CODES, + MISSING_FROM_OMS if
    any) to output_path."""
    wb = Workbook()

    ws_fields = wb.active
    ws_fields.title = "FIELD_CHANGES"
    _write_field_changes_sheet(ws_fields, result.cell_differences)

    ws_diff = wb.create_sheet("DIFFERENCES")
    _write_sheet(ws_diff, oms_sheet, result.changed_rows)

    ws_new = wb.create_sheet("NEW_CODES")
    _write_sheet(ws_new, oms_sheet, result.new_codes)

    if result.missing_from_oms:
        ws_missing = wb.create_sheet("MISSING_FROM_OMS")
        _write_sheet(ws_missing, current_sheet, result.missing_from_oms)

    output_path = Path(output_path)
    wb.save(str(output_path))
    logger.info("Wrote comparison result to %s", output_path)
    return output_path


def default_output_filename():
    date_str = datetime.now().strftime("%Y-%m-%d")
    return f"Garg_Dental_Comparison_{date_str}.xlsx"


def _mapped_headers(mappings):
    headers = ["Code"]
    for m in mappings:
        headers.append(f"{m.current_column} (Current)")
        headers.append(f"{m.latest_column} (Latest)")
    headers.append("Changed Fields")
    return headers


def _write_mapped_rows(ws, mappings, rows, row_fill=None):
    """Writes one row per MappedProductRow. `row_fill`, if given, is applied
    to every cell in the row (Added/Removed sheets); otherwise each row gets
    light green only if it has changes, with individual changed cells
    yellow and missing-value cells orange overriding it - exactly the
    precedence described in the CONDITIONAL FORMATTING spec."""
    ncols = len(_mapped_headers(mappings))
    for row in rows:
        dest_row = ws.max_row + 1
        ws.cell(row=dest_row, column=1, value=row.code)
        col = 2
        for f in row.fields:
            current_cell = ws.cell(row=dest_row, column=col, value=f.current_value)
            latest_cell = ws.cell(row=dest_row, column=col + 1, value=f.latest_value)
            if row_fill is None:
                if f.missing:
                    current_cell.fill = _ORANGE
                    latest_cell.fill = _ORANGE
                elif f.changed:
                    current_cell.fill = _YELLOW
                    latest_cell.fill = _YELLOW
            col += 2
        ws.cell(row=dest_row, column=col, value=", ".join(row.changed_field_labels))

        if row_fill is not None:
            for c in range(1, ncols + 1):
                ws.cell(row=dest_row, column=c).fill = row_fill
        elif row.changed_field_labels:
            for c in range(1, ncols + 1):
                cell = ws.cell(row=dest_row, column=c)
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = _LIGHT_GREEN


def _write_mapped_sheet(ws, mappings, rows, row_fill=None):
    ws.append(_mapped_headers(mappings))
    _write_mapped_rows(ws, mappings, rows, row_fill=row_fill)
    ncols = len(_mapped_headers(mappings))
    style_header_row(ws, ncols)
    autofit_columns(ws, ncols)


def _write_mapped_summary_sheet(ws, result, current_filename, latest_filename):
    ws.append(["Metric", "Value"])
    rows = [
        ("Products Compared", result.total_compared),
        ("Rows Changed", result.total_changed),
        ("Rows Unchanged", result.total_unchanged),
        ("Added Products", result.total_added),
        ("Removed Products", result.total_removed),
        ("", ""),
        ("Differences By Column", ""),
    ]
    change_counts = {m.current_column: 0 for m in result.mappings}
    for row in result.rows:
        for label in row.changed_field_labels:
            change_counts[label] = change_counts.get(label, 0) + 1
    for mapping in result.mappings:
        rows.append((f"  {mapping.current_column} / {mapping.latest_column}", change_counts[mapping.current_column]))
    rows.extend(
        [
            ("", ""),
            ("Comparison Date", result.compared_at.strftime("%Y-%m-%d %H:%M")),
            ("Current File", current_filename or ""),
            ("Latest File", latest_filename or ""),
        ]
    )
    for label, value in rows:
        ws.append([label, value])
    style_header_row(ws, 2)
    autofit_columns(ws, 2)


def write_mapped_result(output_path, result, current_filename=None, latest_filename=None):
    """Write the Column Comparison Report Mode workbook: SUMMARY,
    DIFFERENCES (changed rows only), ADDED_PRODUCTS, REMOVED_PRODUCTS, and
    UNCHANGED_PRODUCTS. Used when 2+ columns are mapped (or 1 mapped pair
    with differently-named columns) - see compare_mapped()."""
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "SUMMARY"
    _write_mapped_summary_sheet(ws_summary, result, current_filename, latest_filename)

    changed_rows = [r for r in result.rows if r.changed_field_labels]
    unchanged_rows = [r for r in result.rows if not r.changed_field_labels]

    ws_diff = wb.create_sheet("DIFFERENCES")
    _write_mapped_sheet(ws_diff, result.mappings, changed_rows)

    ws_added = wb.create_sheet("ADDED_PRODUCTS")
    _write_mapped_sheet(ws_added, result.mappings, result.added, row_fill=_LIGHT_BLUE)

    ws_removed = wb.create_sheet("REMOVED_PRODUCTS")
    _write_mapped_sheet(ws_removed, result.mappings, result.removed, row_fill=_LIGHT_RED)

    ws_unchanged = wb.create_sheet("UNCHANGED_PRODUCTS")
    _write_mapped_sheet(ws_unchanged, result.mappings, unchanged_rows)

    output_path = Path(output_path)
    wb.save(str(output_path))
    logger.info("Wrote mapped comparison result to %s", output_path)
    return output_path


def default_mapped_output_filename():
    date_str = datetime.now().strftime("%Y-%m-%d")
    return f"Garg_Dental_Column_Comparison_{date_str}.xlsx"
