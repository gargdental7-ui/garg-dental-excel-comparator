"""Writes the Collection Priority Analyzer output workbook."""
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from .report_style import autofit_columns, style_header_row

PRIORITY_HEADER = [
    "Priority Rank",
    "Priority Category",
    "Customer Name",
    "Total Outstanding",
    "Open Invoice Count",
    "Maximum Days Overdue",
    "Average Days Overdue",
    "Oldest Due Date",
    "Salesperson",
]


def _write_priority_sheet(ws, customers):
    ws.append(PRIORITY_HEADER)
    for rank, cust in enumerate(customers, start=1):
        ws.append(
            [
                rank,
                cust.priority,
                cust.customer,
                round(cust.total_outstanding, 2),
                cust.invoice_count,
                round(cust.max_days_overdue, 1),
                round(cust.avg_days_overdue, 1),
                cust.oldest_due_date,
                cust.salesperson or "",
            ]
        )
    style_header_row(ws, len(PRIORITY_HEADER))
    autofit_columns(ws, len(PRIORITY_HEADER))


def _write_invoice_details_sheet(ws, headers, rows):
    ws.append(list(headers))
    for row in rows:
        ws.append([row.get(h) for h in headers])
    style_header_row(ws, len(headers))
    autofit_columns(ws, len(headers))


def _write_summary_sheet(ws, result):
    t = result.thresholds
    ws.append(["Metric", "Value"])
    for label, value in [
        ("Total Outstanding", round(result.total_outstanding, 2)),
        ("Total Accounts", result.total_customers),
        ("Critical Accounts", result.critical_count),
        ("High Priority Accounts", result.high_count),
        ("Critical Amount", round(result.critical_amount, 2)),
        ("High Priority Amount", round(result.high_priority_amount, 2)),
        ("Amount Over 30 Days", round(result.amount_over_30, 2)),
        ("Amount Over 60 Days", round(result.amount_over_60, 2)),
        ("Amount Over 90 Days", round(result.amount_over_90, 2)),
        ("", ""),
        ("Rule: Critical Overdue Days", t.critical_days),
        ("Rule: High Overdue Days", t.high_days),
        ("Rule: Medium Overdue Days", t.medium_days),
        ("Rule: Critical Amount Threshold", t.critical_amount),
        ("Rule: High Amount Threshold", t.high_amount),
        ("Rule: Many-Invoices Bump Count", t.many_invoices_count),
    ]:
        ws.append([label, value])
    style_header_row(ws, 2)
    autofit_columns(ws, 2)


def write_collection_report(output_path, result):
    """Write COLLECTION_PRIORITY (+ INVOICE_DETAILS, + SUMMARY) to
    output_path. Never touches the uploaded source file."""
    wb = Workbook()

    ws_priority = wb.active
    ws_priority.title = "COLLECTION_PRIORITY"
    _write_priority_sheet(ws_priority, result.customers)

    ws_invoices = wb.create_sheet("INVOICE_DETAILS")
    _write_invoice_details_sheet(ws_invoices, result.headers, result.rows)

    ws_summary = wb.create_sheet("SUMMARY")
    _write_summary_sheet(ws_summary, result)

    output_path = Path(output_path)
    wb.save(str(output_path))
    return output_path


def default_output_filename():
    date_str = datetime.now().strftime("%Y-%m-%d")
    return f"Garg_Dental_Collection_Priority_{date_str}.xlsx"
