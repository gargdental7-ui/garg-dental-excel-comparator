import pytest

from app import collection_analyzer, collection_writer
from app.collection_analyzer import CollectionColumnMapping, CollectionThresholds
from app.exceptions import MissingColumnMappingError, NoDataRowsError

HEADERS = ["Customer", "Amount", "Days Overdue", "Due Date", "Salesperson", "Invoice No"]


def _mapping(**overrides):
    base = dict(
        customer="Customer",
        amount="Amount",
        days_overdue="Days Overdue",
        due_date="Due Date",
        salesperson="Salesperson",
        invoice_number="Invoice No",
    )
    base.update(overrides)
    return CollectionColumnMapping(**base)


def _row(customer, amount, days_overdue=0, due_date=None, salesperson=None, invoice_no=None):
    return {
        "Customer": customer,
        "Amount": amount,
        "Days Overdue": days_overdue,
        "Due Date": due_date,
        "Salesperson": salesperson,
        "Invoice No": invoice_no,
    }


def test_multiple_invoices_aggregate_correctly_by_customer():
    rows = [
        _row("Acme Dental", 1000, 10),
        _row("Acme Dental", 2000, 40),
        _row("Best Smiles", 500, 5),
    ]
    result = collection_analyzer.analyze_collections(HEADERS, rows, _mapping())
    acme = next(c for c in result.customers if c.customer == "Acme Dental")
    assert acme.total_outstanding == 3000
    assert acme.invoice_count == 2
    assert acme.max_days_overdue == 40
    assert acme.avg_days_overdue == 25


def test_outstanding_totals_are_correct():
    rows = [_row("A", 100), _row("B", 250), _row("A", 50)]
    result = collection_analyzer.analyze_collections(HEADERS, rows, _mapping())
    assert result.total_outstanding == 400
    assert result.total_customers == 2


def test_days_overdue_computed_from_due_date_when_column_absent():
    from datetime import date, timedelta

    overdue_date = date.today() - timedelta(days=50)
    rows = [_row("A", 1000, days_overdue=None, due_date=overdue_date.isoformat())]
    mapping = _mapping(days_overdue=None)
    result = collection_analyzer.analyze_collections(HEADERS, rows, mapping)
    assert result.customers[0].max_days_overdue == 50


def test_priority_ranking_is_deterministic():
    thresholds = CollectionThresholds(
        critical_days=90, high_days=60, medium_days=30, critical_amount=100000, high_amount=50000
    )
    rows = [
        _row("Critical Co", 5000, 95),
        _row("High Co", 5000, 65),
        _row("Medium Co", 5000, 35),
        _row("Normal Co", 5000, 5),
    ]
    result = collection_analyzer.analyze_collections(HEADERS, rows, _mapping(), thresholds)
    by_name = {c.customer: c.priority for c in result.customers}
    assert by_name == {
        "Critical Co": "CRITICAL",
        "High Co": "HIGH",
        "Medium Co": "MEDIUM",
        "Normal Co": "NORMAL",
    }
    # Sorted highest priority first.
    assert [c.customer for c in result.customers] == ["Critical Co", "High Co", "Medium Co", "Normal Co"]


def test_many_invoices_bump_priority_one_level():
    thresholds = CollectionThresholds(many_invoices_count=3)
    rows = [_row("Frequent Co", 100, 5) for _ in range(3)]
    result = collection_analyzer.analyze_collections(HEADERS, rows, _mapping(), thresholds)
    # Amount/days alone would be NORMAL; 3 invoices bumps it to MEDIUM.
    assert result.customers[0].priority == "MEDIUM"


def test_blank_amounts_are_handled_safely():
    rows = [_row("A", None), _row("A", "")]
    result = collection_analyzer.analyze_collections(HEADERS, rows, _mapping())
    assert result.customers[0].total_outstanding == 0
    assert result.customers[0].invoice_count == 2


def test_negative_credit_balances_handled_explicitly():
    rows = [_row("Credit Customer", -500, 120)]
    result = collection_analyzer.analyze_collections(HEADERS, rows, _mapping())
    cust = result.customers[0]
    assert cust.total_outstanding == -500
    assert cust.priority == "NORMAL"  # nothing to collect despite high days-overdue
    assert result.critical_count == 0
    assert result.critical_amount == 0


def test_missing_required_columns_raise_clear_error():
    with pytest.raises(MissingColumnMappingError):
        collection_analyzer.analyze_collections(HEADERS, [_row("A", 1)], CollectionColumnMapping(customer="", amount=""))


def test_no_data_rows_raises():
    with pytest.raises(NoDataRowsError):
        collection_analyzer.analyze_collections(HEADERS, [], _mapping())


def test_blank_customer_rows_are_skipped():
    rows = [_row("", 100), _row("Acme Dental", 200)]
    result = collection_analyzer.analyze_collections(HEADERS, rows, _mapping())
    assert result.total_customers == 1


def test_export_workbook_contains_expected_sheets_and_preserves_invoice_rows(tmp_path):
    rows = [_row("Acme Dental", 1000, 10, invoice_no="INV-1"), _row("Acme Dental", 2000, 95, invoice_no="INV-2")]
    result = collection_analyzer.analyze_collections(HEADERS, rows, _mapping())

    output_path = tmp_path / "collection.xlsx"
    collection_writer.write_collection_report(output_path, result)

    import openpyxl

    wb = openpyxl.load_workbook(output_path)
    assert wb.sheetnames == ["COLLECTION_PRIORITY", "INVOICE_DETAILS", "SUMMARY"]

    ws_priority = wb["COLLECTION_PRIORITY"]
    assert ws_priority.cell(row=2, column=3).value == "Acme Dental"

    ws_invoices = wb["INVOICE_DETAILS"]
    invoice_ids = [ws_invoices.cell(row=r, column=6).value for r in range(2, ws_invoices.max_row + 1)]
    assert set(invoice_ids) == {"INV-1", "INV-2"}
