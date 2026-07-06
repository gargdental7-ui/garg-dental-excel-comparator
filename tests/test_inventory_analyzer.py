import pytest

from app import inventory_analyzer, inventory_writer
from app.inventory_analyzer import (
    FAST_MOVING,
    NEGATIVE_STOCK,
    NO_MOVEMENT,
    NORMAL_MOVING,
    OUT_OF_STOCK,
    SLOW_MOVING,
    InventoryColumnMapping,
    MovementThresholds,
    classify_movement,
)
from app.exceptions import MissingColumnMappingError, NoDataRowsError

HEADERS = ["Code", "Description", "Unit", "Opening", "Received", "Delivered", "Balance", "Unit Cost"]


def _mapping(**overrides):
    base = dict(
        code="Code",
        description="Description",
        unit="Unit",
        opening="Opening",
        received="Received",
        delivered="Delivered",
        balance="Balance",
    )
    base.update(overrides)
    return InventoryColumnMapping(**base)


def _row(code, opening, received, delivered, balance, unit_cost=None, description="Desc", unit="EA"):
    return {
        "Code": code,
        "Description": description,
        "Unit": unit,
        "Opening": opening,
        "Received": received,
        "Delivered": delivered,
        "Balance": balance,
        "Unit Cost": unit_cost,
    }


def test_delivered_zero_and_balance_positive_is_no_movement():
    classification, ratio = classify_movement(10, 0, 0, 10, MovementThresholds())
    assert classification == NO_MOVEMENT
    assert ratio == 0.0


def test_negative_balance_is_negative_stock():
    classification, _ = classify_movement(10, 0, 5, -5, MovementThresholds())
    assert classification == NEGATIVE_STOCK


def test_balance_zero_is_out_of_stock():
    classification, _ = classify_movement(10, 0, 10, 0, MovementThresholds())
    assert classification == OUT_OF_STOCK


def test_movement_ratio_calculates_correctly():
    classification, ratio = classify_movement(10, 10, 16, 4, MovementThresholds())
    assert ratio == pytest.approx(0.8)
    assert classification == FAST_MOVING


def test_division_by_zero_is_safe():
    classification, ratio = classify_movement(0, 0, 5, 5, MovementThresholds())
    assert ratio == 0.0
    assert classification == SLOW_MOVING


def test_threshold_configuration_changes_classification():
    loose = MovementThresholds(fast_ratio=0.9, normal_ratio=0.1)
    classification, ratio = classify_movement(10, 0, 5, 5, loose)
    assert ratio == pytest.approx(0.5)
    assert classification == NORMAL_MOVING

    strict = MovementThresholds(fast_ratio=0.4, normal_ratio=0.1)
    classification2, _ = classify_movement(10, 0, 5, 5, strict)
    assert classification2 == FAST_MOVING


def test_complete_product_information_is_preserved():
    rows = [_row("A1", 10, 10, 15, 5)]
    result = inventory_analyzer.analyze_inventory(HEADERS, rows, [3], _mapping())
    product = result.products[0]
    assert product.code == "A1"
    assert product.description == "Desc"
    assert product.unit == "EA"
    assert product.opening == 10
    assert product.received == 10
    assert product.delivered == 15
    assert product.balance == 5


def test_duplicate_codes_are_detected():
    rows = [_row("A1", 10, 0, 5, 5), _row("A1", 20, 0, 5, 15)]
    result = inventory_analyzer.analyze_inventory(HEADERS, rows, [3, 4], _mapping())
    assert result.products == []
    reasons = {(exc.code, exc.reason) for exc in result.exceptions}
    assert reasons == {("A1", "Duplicate Code")}


def test_missing_code_is_an_exception():
    rows = [_row("", 10, 0, 5, 5)]
    result = inventory_analyzer.analyze_inventory(HEADERS, rows, [3], _mapping())
    assert result.products == []
    assert result.exceptions[0].reason == "Missing Code"


def test_invalid_numeric_value_is_an_exception():
    rows = [_row("A1", "not-a-number", 0, 5, 5)]
    result = inventory_analyzer.analyze_inventory(HEADERS, rows, [3], _mapping())
    assert result.products == []
    assert result.exceptions[0].reason == "Invalid numeric values"


def test_no_group_rows_are_silently_skipped():
    rows = [_row("No Group", None, None, None, None), _row("A1", 10, 0, 5, 5)]
    result = inventory_analyzer.analyze_inventory(HEADERS, rows, [3, 4], _mapping())
    assert len(result.products) == 1
    assert result.products[0].code == "A1"
    assert result.exceptions == []


def test_two_row_header_workbook_feeds_analyzer_end_to_end(tmp_path):
    from app.generic_excel import load_generic_sheet
    from tests.factories import write_two_row_workbook
    import openpyxl

    path = tmp_path / "stock.xlsx"
    write_two_row_workbook(path, [["AD00668", "AD AIR MOTER", "EA", 22, 2, 20, 4]])
    wb = openpyxl.load_workbook(path, data_only=True)
    data = load_generic_sheet(wb, wb.sheetnames[0], "Test File")

    result = inventory_analyzer.analyze_inventory(
        data.headers, data.rows, data.row_excel_indexes, _mapping()
    )
    assert len(result.products) == 1
    assert result.products[0].code == "AD00668"


def test_high_value_analysis_only_runs_when_value_data_exists():
    rows = [_row("A1", 10, 0, 0, 5, unit_cost=100)]
    result_with_value = inventory_analyzer.analyze_inventory(
        HEADERS, rows, [3], _mapping(unit_cost="Unit Cost")
    )
    assert result_with_value.has_value_data is True
    assert result_with_value.total_inventory_value == 500
    assert result_with_value.products[0].inventory_value == 500

    result_without_value = inventory_analyzer.analyze_inventory(HEADERS, rows, [3], _mapping())
    assert result_without_value.has_value_data is False
    assert result_without_value.total_inventory_value is None
    assert result_without_value.products[0].inventory_value is None


def test_missing_required_mapping_raises():
    with pytest.raises(MissingColumnMappingError):
        inventory_analyzer.analyze_inventory(
            HEADERS,
            [_row("A1", 1, 1, 1, 1)],
            [3],
            InventoryColumnMapping(code="", opening="", received="", delivered="", balance=""),
        )


def test_no_data_rows_raises():
    with pytest.raises(NoDataRowsError):
        inventory_analyzer.analyze_inventory(HEADERS, [], [], _mapping())


def test_export_workbook_contains_correct_sheets(tmp_path):
    rows = [
        _row("FAST1", 10, 10, 16, 4, unit_cost=50),   # FAST MOVING
        _row("NOMOVE1", 10, 0, 0, 10, unit_cost=200),  # NO MOVEMENT
        _row("SLOW1", 10, 0, 1, 9, unit_cost=10),      # SLOW MOVING
        _row("NEG1", 10, 0, 5, -1),                    # NEGATIVE STOCK
        _row("OUT1", 10, 0, 10, 0),                    # OUT OF STOCK
    ]
    result = inventory_analyzer.analyze_inventory(
        HEADERS, rows, list(range(3, 3 + len(rows))), _mapping(unit_cost="Unit Cost")
    )

    output_path = tmp_path / "inventory.xlsx"
    inventory_writer.write_inventory_report(output_path, result)

    import openpyxl

    wb = openpyxl.load_workbook(output_path)
    assert wb.sheetnames == [
        "INVENTORY_ANALYSIS",
        "NO_MOVEMENT",
        "SLOW_MOVING",
        "FAST_MOVING",
        "STOCK_EXCEPTIONS",
        "HIGH_VALUE_RISK",
    ]
    assert wb["INVENTORY_ANALYSIS"].max_row == 1 + len(rows)
    assert wb["NO_MOVEMENT"].max_row == 2
    assert wb["FAST_MOVING"].max_row == 2

    risk_codes = [wb["HIGH_VALUE_RISK"].cell(row=r, column=1).value for r in range(2, wb["HIGH_VALUE_RISK"].max_row + 1)]
    assert set(risk_codes) == {"NOMOVE1", "SLOW1"}


def test_export_hides_value_sheet_when_no_value_data(tmp_path):
    rows = [_row("A1", 10, 0, 0, 10)]
    result = inventory_analyzer.analyze_inventory(HEADERS, rows, [3], _mapping())

    output_path = tmp_path / "inventory.xlsx"
    inventory_writer.write_inventory_report(output_path, result)

    import openpyxl

    wb = openpyxl.load_workbook(output_path)
    assert "HIGH_VALUE_RISK" not in wb.sheetnames
