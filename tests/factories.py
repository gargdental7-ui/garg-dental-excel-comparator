"""Helpers for building sample .xlsx workbooks in tests."""
import openpyxl

HEADER_ROW_1 = ["Code", "Description", "Unit", "Opening", "Received", "Delivered", "Balance"]
HEADER_ROW_2 = [None, None, None, "Qty", "Qty", "Qty", "Qty"]


def write_two_row_workbook(path, data_rows):
    """Write a workbook using the known Garg Dental two-row header layout."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADER_ROW_1)
    ws.append(HEADER_ROW_2)
    for row in data_rows:
        ws.append(row)
    wb.save(path)
    return path


def write_simple_workbook(path, headers, data_rows):
    """Write a workbook with an ordinary single-row header."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in data_rows:
        ws.append(row)
    wb.save(path)
    return path


def write_titled_two_row_workbook(path, data_rows):
    """Write a workbook with a report title block preceding the known
    Garg Dental two-row header layout, as real OMS exports do."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["FY : 01/04/2082-32/03/2083"])
    ws.append([])
    ws.append(["Garg Dental Pvt. Ltd. F/Y 82-83"])
    ws.append([])
    ws.append(HEADER_ROW_1)
    ws.append(HEADER_ROW_2)
    for row in data_rows:
        ws.append(row)
    wb.save(path)
    return path
