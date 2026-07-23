import openpyxl

from app import comparator, workbook_writer
from app.comparator import ColumnMapping
from app.workbook_reader import load_sheet

from .factories import write_simple_workbook


def _load(tmp_path, name, headers, rows):
    path = tmp_path / name
    write_simple_workbook(path, headers, rows)
    return load_sheet(path, name)


def _build_result(tmp_path):
    current = _load(
        tmp_path,
        "current.xlsx",
        ["Code", "Balance", "Price"],
        [["A1", 10, 100], ["A2", 5, 50], ["A3", 1, 1]],
    )
    oms = _load(
        tmp_path,
        "oms.xlsx",
        ["Code", "Available Stock", "Selling Price"],
        [["A1", 10, 120], ["A2", 5, 50], ["A4", 9, 90]],
    )
    mappings = [
        ColumnMapping(current_column="Balance", latest_column="Available Stock"),
        ColumnMapping(current_column="Price", latest_column="Selling Price"),
    ]
    return comparator.compare_mapped(current, oms, mappings)


def test_write_mapped_result_produces_all_sheets(tmp_path):
    result = _build_result(tmp_path)
    out_path = tmp_path / "out.xlsx"
    workbook_writer.write_mapped_result(out_path, result, "current.xlsx", "oms.xlsx")

    wb = openpyxl.load_workbook(out_path)
    assert wb.sheetnames == ["SUMMARY", "DIFFERENCES", "ADDED_PRODUCTS", "REMOVED_PRODUCTS", "UNCHANGED_PRODUCTS"]


def test_differences_sheet_has_changed_fields_column_and_yellow_fill(tmp_path):
    result = _build_result(tmp_path)
    out_path = tmp_path / "out.xlsx"
    workbook_writer.write_mapped_result(out_path, result, "current.xlsx", "oms.xlsx")

    wb = openpyxl.load_workbook(out_path)
    ws = wb["DIFFERENCES"]
    header = [c.value for c in ws[1]]
    assert header == ["Code", "Balance (Current)", "Available Stock (Latest)", "Price (Current)", "Selling Price (Latest)", "Changed Fields"]

    # Only A1 changed (Price 100 -> 120); A2 unchanged should not appear here.
    data_rows = list(ws.iter_rows(min_row=2, values_only=False))
    assert len(data_rows) == 1
    row = data_rows[0]
    assert row[0].value == "A1"
    assert row[5].value == "Price"
    # Price (Current)=100, Price (Latest)=120 are the changed cells -> yellow
    assert row[3].fill.fgColor.rgb == "00FFFF00"
    assert row[4].fill.fgColor.rgb == "00FFFF00"
    # Balance cells didn't change individually, but the row fill (light green)
    # should still apply since the row has a change.
    assert row[1].fill.fgColor.rgb == "00CCFFCC"


def test_added_products_sheet_has_blue_fill_and_missing_current_values(tmp_path):
    result = _build_result(tmp_path)
    out_path = tmp_path / "out.xlsx"
    workbook_writer.write_mapped_result(out_path, result, "current.xlsx", "oms.xlsx")

    wb = openpyxl.load_workbook(out_path)
    ws = wb["ADDED_PRODUCTS"]
    rows = list(ws.iter_rows(min_row=2, values_only=False))
    assert len(rows) == 1
    row = rows[0]
    assert row[0].value == "A4"
    assert row[1].value is None  # Balance (Current) - product didn't exist in current file
    assert row[2].value == 9  # Available Stock (Latest)
    assert row[0].fill.fgColor.rgb == "00BDD7EE"


def test_removed_products_sheet_has_red_fill(tmp_path):
    result = _build_result(tmp_path)
    out_path = tmp_path / "out.xlsx"
    workbook_writer.write_mapped_result(out_path, result, "current.xlsx", "oms.xlsx")

    wb = openpyxl.load_workbook(out_path)
    ws = wb["REMOVED_PRODUCTS"]
    rows = list(ws.iter_rows(min_row=2, values_only=False))
    assert len(rows) == 1
    assert rows[0][0].value == "A3"
    assert rows[0][0].fill.fgColor.rgb == "00FFC7CE"


def test_unchanged_products_sheet_has_no_special_fill(tmp_path):
    result = _build_result(tmp_path)
    out_path = tmp_path / "out.xlsx"
    workbook_writer.write_mapped_result(out_path, result, "current.xlsx", "oms.xlsx")

    wb = openpyxl.load_workbook(out_path)
    ws = wb["UNCHANGED_PRODUCTS"]
    rows = list(ws.iter_rows(min_row=2, values_only=False))
    assert len(rows) == 1
    assert rows[0][0].value == "A2"
    assert rows[0][0].fill.fgColor.rgb in (None, "00000000")


def test_summary_sheet_reports_counts_and_files(tmp_path):
    result = _build_result(tmp_path)
    out_path = tmp_path / "out.xlsx"
    workbook_writer.write_mapped_result(out_path, result, "current_report.xlsx", "latest_report.xlsx")

    wb = openpyxl.load_workbook(out_path)
    ws = wb["SUMMARY"]
    values = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2, values_only=False) if row[0].value}
    assert values["Products Compared"] == 2
    assert values["Rows Changed"] == 1
    assert values["Rows Unchanged"] == 1
    assert values["Added Products"] == 1
    assert values["Removed Products"] == 1
    assert values["Current File"] == "current_report.xlsx"
    assert values["Latest File"] == "latest_report.xlsx"


def test_default_mapped_output_filename_format():
    name = workbook_writer.default_mapped_output_filename()
    assert name.startswith("Garg_Dental_Column_Comparison_")
    assert name.endswith(".xlsx")
