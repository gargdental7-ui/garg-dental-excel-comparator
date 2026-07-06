"""End-to-end: build two realistic Garg Dental workbooks, compare, write the
result, reopen it, and verify only the correct complete OMS rows survive."""
import openpyxl

from app import comparator, workbook_writer
from app.workbook_reader import load_sheet

from .factories import write_two_row_workbook


def test_end_to_end_comparison(tmp_path):
    current_rows = [
        ["AD00668", "AD AIR MOTER", "EA", 22, 2, 20, 4],
        ["AI00001", "AIR PIPE FOR DENTAL TROLLEY (5MTR)", "SET", None, 2, None, 2],
        ["AN00002", "ANLOMETIC VOLTAGE", "EA", 1, None, None, 1],
        ["HFILE", "H FILE 25MM 6, 8, 10,15", "PCS", None, 16, 24, -8],
    ]
    oms_rows = [
        ["AD00668", "AD AIR MOTER", "EA", 22, 2, 20, 4],  # same Balance -> excluded
        ["AI00001", "AIR PIPE FOR DENTAL TROLLEY (5MTR)", "SET", None, 2, None, 5],  # different -> included
        ["AN00002", "ANLOMETIC VOLTAGE", "EA", 1, None, None, 1],  # same -> excluded
        ["HFILE", "H FILE 25MM 6, 8, 10,15", "PCS", None, 16, 24, -9],  # different -> included
    ]

    current_path = tmp_path / "current.xlsx"
    oms_path = tmp_path / "oms.xlsx"
    write_two_row_workbook(current_path, current_rows)
    write_two_row_workbook(oms_path, oms_rows)

    current = load_sheet(current_path, "Current File")
    oms = load_sheet(oms_path, "OMS File")
    result = comparator.compare(current, oms)

    assert result.total_compared == 4
    assert result.total_differences == 2
    assert result.new_codes == []
    assert result.missing_from_oms == []

    output_path = tmp_path / "result.xlsx"
    workbook_writer.write_result(output_path, current, oms, result)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["DIFFERENCES"]
    data_rows = [[c.value for c in ws[r]] for r in range(3, ws.max_row + 1)]

    assert len(data_rows) == 2
    assert ["AI00001", "AIR PIPE FOR DENTAL TROLLEY (5MTR)", "SET", None, 2, None, 5] in data_rows
    assert ["HFILE", "H FILE 25MM 6, 8, 10,15", "PCS", None, 16, 24, -9] in data_rows

    # Unchanged rows must not appear.
    codes_in_output = {row[0] for row in data_rows}
    assert "AD00668" not in codes_in_output
    assert "AN00002" not in codes_in_output


def test_end_to_end_with_new_and_missing_codes(tmp_path):
    current_rows = [
        ["A1", "Item A", "EA", 1, 1, 1, 1],
        ["A3", "Item C (current only)", "EA", 1, 1, 1, 1],
    ]
    oms_rows = [
        ["A1", "Item A", "EA", 1, 1, 1, 2],
        ["A2", "Item B (oms only)", "EA", 1, 1, 1, 1],
    ]

    current_path = tmp_path / "current.xlsx"
    oms_path = tmp_path / "oms.xlsx"
    write_two_row_workbook(current_path, current_rows)
    write_two_row_workbook(oms_path, oms_rows)

    current = load_sheet(current_path, "Current File")
    oms = load_sheet(oms_path, "OMS File")
    result = comparator.compare(current, oms)

    output_path = tmp_path / "result.xlsx"
    workbook_writer.write_result(output_path, current, oms, result)

    wb = openpyxl.load_workbook(output_path)
    diff_rows = [[c.value for c in wb["DIFFERENCES"][r]] for r in range(3, wb["DIFFERENCES"].max_row + 1)]
    new_rows = [[c.value for c in wb["NEW_CODES"][r]] for r in range(3, wb["NEW_CODES"].max_row + 1)]
    missing_rows = [
        [c.value for c in wb["MISSING_FROM_OMS"][r]] for r in range(3, wb["MISSING_FROM_OMS"].max_row + 1)
    ]

    assert ["A1", "Item A", "EA", 1, 1, 1, 2] in diff_rows
    assert ["A2", "Item B (oms only)", "EA", 1, 1, 1, 1] in new_rows
    assert ["A3", "Item C (current only)", "EA", 1, 1, 1, 1] in missing_rows
