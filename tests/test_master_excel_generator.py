from openpyxl import load_workbook

from app.master_excel_generator import build_master_excel


def test_build_master_excel_writes_recognized_headers(tmp_path):
    out = build_master_excel(
        tmp_path / "master.xlsx",
        [{"product_name": "Widget A", "price": 100, "code": "W1", "brand": "Acme"}],
    )
    wb = load_workbook(out)
    ws = wb["PRODUCTS"]
    headers = [cell.value for cell in ws[1]]
    assert headers == [
        "Product Name",
        "Price",
        "Code",
        "Description",
        "Brand",
        "Model",
        "Origin",
        "Category",
        "Warranty",
        "MRP",
    ]


def test_build_master_excel_fills_missing_fields_blank(tmp_path):
    out = build_master_excel(tmp_path / "master.xlsx", [{"product_name": "Widget A"}])
    wb = load_workbook(out)
    ws = wb["PRODUCTS"]
    row = [cell.value for cell in ws[2]]
    assert row[0] == "Widget A"
    assert not row[2]  # code, blank rather than missing (openpyxl stores "" as None)


def test_build_master_excel_handles_empty_product_list(tmp_path):
    out = build_master_excel(tmp_path / "master.xlsx", [])
    wb = load_workbook(out)
    ws = wb["PRODUCTS"]
    assert ws.max_row == 1  # header row only
