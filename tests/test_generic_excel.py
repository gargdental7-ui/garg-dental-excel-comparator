import openpyxl
import pytest

from app import generic_excel
from app.exceptions import GenericHeaderDetectionError


def _wb_from_rows(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    return wb


def test_simple_single_row_header_detected():
    wb = _wb_from_rows(
        [
            ["Customer", "Outstanding Amount", "Days Overdue"],
            ["Acme Dental", 5000, 45],
            ["Best Smiles", 12000, 10],
        ]
    )
    data = generic_excel.load_generic_sheet(wb, "Sheet", "Test File")
    assert data.headers == ["Customer", "Outstanding Amount", "Days Overdue"]
    assert data.header_row_count == 1
    assert len(data.rows) == 2
    assert data.rows[0]["Customer"] == "Acme Dental"


def test_header_after_report_title_block_detected():
    wb = _wb_from_rows(
        [
            ["Garg Dental Pvt. Ltd. - Outstanding Report"],
            ["Date From: 01/04/2082 To: 32/03/2083"],
            [],
            ["Customer", "Outstanding Amount", "Due Date"],
            ["Acme Dental", 5000, "2026-01-01"],
        ]
    )
    data = generic_excel.load_generic_sheet(wb, "Sheet", "Test File")
    assert data.headers == ["Customer", "Outstanding Amount", "Due Date"]
    assert data.header_row_start == 4
    assert len(data.rows) == 1


def test_known_garg_two_row_header_detected():
    wb = _wb_from_rows(
        [
            ["FY : 01/04/2082-32/03/2083"],
            [],
            ["Code", "Description", "Unit", "Opening", "Received", "Delivered", "Balance"],
            [None, None, None, "Qty", "Qty", "Qty", "Qty"],
            ["AD00668", "AD AIR MOTER", "EA", 22, 2, 20, 4],
        ]
    )
    data = generic_excel.load_generic_sheet(wb, "Sheet", "Test File")
    assert data.headers == ["Code", "Description", "Unit", "Opening", "Received", "Delivered", "Balance"]
    assert data.header_row_count == 2
    assert len(data.rows) == 1
    assert data.rows[0]["Code"] == "AD00668"


def test_blank_rows_between_header_and_data_are_skipped():
    wb = _wb_from_rows(
        [
            ["Customer", "Outstanding Amount"],
            [None, None],
            ["Acme Dental", 5000],
        ]
    )
    data = generic_excel.load_generic_sheet(wb, "Sheet", "Test File")
    assert len(data.rows) == 1


def test_duplicate_headers_are_deduped():
    wb = _wb_from_rows(
        [
            ["Customer", "Amount", "Amount"],
            ["Acme Dental", 100, 200],
        ]
    )
    data = generic_excel.load_generic_sheet(wb, "Sheet", "Test File")
    assert data.headers == ["Customer", "Amount", "Amount (2)"]


def test_trailing_blank_header_column_is_trimmed():
    wb = _wb_from_rows(
        [
            ["Customer", "Amount", None],
            ["Acme Dental", 100, "ignored"],
        ]
    )
    data = generic_excel.load_generic_sheet(wb, "Sheet", "Test File")
    assert data.headers == ["Customer", "Amount"]


def test_no_header_row_raises():
    wb = _wb_from_rows([[None, None], [None, None]])
    with pytest.raises(GenericHeaderDetectionError):
        generic_excel.load_generic_sheet(wb, "Sheet", "Test File")


def test_multiple_sheets_listed():
    wb = _wb_from_rows([["Customer", "Amount"], ["A", 1]])
    wb.create_sheet("Second")
    assert wb.sheetnames == ["Sheet", "Second"]


def test_detect_header_row_index_returns_none_instead_of_raising():
    wb = _wb_from_rows([[None, None], [None, None]])
    assert generic_excel.detect_header_row_index(wb.active) is None


def test_detect_header_row_index_finds_row_1_immediately():
    wb = _wb_from_rows(
        [
            ["Product Name", "Price", "Brand"],
            ["RVG Sensor", 108000, "Woodpecker"],
        ]
    )
    result = generic_excel.detect_header_row_index(wb.active)
    assert result is not None
    row_idx, cleaned, _max_col, header_row_count = result
    assert row_idx == 1
    assert cleaned == ["Product Name", "Price", "Brand"]
    assert header_row_count == 1


def _row_of_blanks(n):
    return [None] * n


def test_detect_header_row_index_scans_forward_for_erp_style_export():
    rows = (
        [["Product Catalog Export"], ["Generated: 2026-07-25"], []]
        + [[] for _ in range(6)]
        + [["Item Name", "Unit Price", "Manufacturer"], ["Autoclave", 50000, "Acme"]]
    )
    wb = _wb_from_rows(rows)
    result = generic_excel.detect_header_row_index(wb.active)
    assert result is not None
    row_idx, cleaned, _max_col, _count = result
    assert row_idx == 10
    assert cleaned == ["Item Name", "Unit Price", "Manufacturer"]


@pytest.mark.parametrize("header_row", [1, 5, 11, 15])
def test_load_generic_sheet_at_row_reads_header_at_any_row(header_row):
    rows = [[] for _ in range(header_row - 1)]
    rows.append(["Product Name", "Price", "Brand", "Model"])
    rows.append(["RVG Sensor", 108000, "Woodpecker", "H1"])
    rows.append(["Portable X-Ray", 175000, "Woodpecker", "AI RAY"])
    wb = _wb_from_rows(rows)

    data = generic_excel.load_generic_sheet_at_row(wb, wb.sheetnames[0], header_row)

    assert data.headers == ["Product Name", "Price", "Brand", "Model"]
    assert data.header_row_start == header_row
    assert data.header_row_count == 1
    assert len(data.rows) == 2
    assert data.rows[0]["Product Name"] == "RVG Sensor"
    assert data.rows[1]["Brand"] == "Woodpecker"


def test_load_generic_sheet_at_row_ignores_two_row_subheader_heuristic():
    # A manually-chosen row is trusted as-is - the very next row is always
    # treated as data, even if it happens to look like a Garg-style
    # sub-header row (contrast with test_known_garg_two_row_header_detected,
    # where the same shape of data is auto-detected as a 2-row header and
    # that sub-header row is consumed instead of becoming a data row).
    wb = _wb_from_rows(
        [
            ["Code", "Description", "Unit", "Opening"],
            [None, None, None, "Qty"],
            ["AD00668", "AD AIR MOTER", "EA", 22],
        ]
    )
    data = generic_excel.load_generic_sheet_at_row(wb, wb.sheetnames[0], 1)
    assert data.header_row_count == 1
    assert len(data.rows) == 2
    assert data.rows[0]["Code"] is None
    assert data.rows[0]["Opening"] == "Qty"
    assert data.rows[1]["Code"] == "AD00668"


def test_load_generic_sheet_at_row_accepts_arbitrary_column_names():
    wb = _wb_from_rows(
        [
            ["Item Description", "Selling Price", "Country of Origin", "Vendor"],
            ["Ultrasonic Scaler", 25000, "Germany", "Dentsply"],
        ]
    )
    data = generic_excel.load_generic_sheet_at_row(wb, wb.sheetnames[0], 1)
    assert data.headers == ["Item Description", "Selling Price", "Country of Origin", "Vendor"]
    assert data.rows[0]["Selling Price"] == 25000


def test_read_preview_rows_returns_raw_rows_with_row_numbers():
    wb = _wb_from_rows(
        [
            ["Product Catalog"],
            [],
            ["Product Name", "Price"],
            ["RVG Sensor", 108000],
        ]
    )
    preview = generic_excel.read_preview_rows(wb.active, max_rows=20)
    assert [p["row"] for p in preview] == [1, 2, 3, 4]
    # Rows are padded out to the sheet's full column count (2, from the
    # widest row), not just each row's own length.
    assert preview[0]["values"] == ["Product Catalog", ""]
    assert preview[2]["values"] == ["Product Name", "Price"]
    assert preview[3]["values"] == ["RVG Sensor", "108000"]


def test_read_preview_rows_caps_at_max_rows():
    wb = _wb_from_rows([[f"Row {i}"] for i in range(1, 31)])
    preview = generic_excel.read_preview_rows(wb.active, max_rows=20)
    assert len(preview) == 20
    assert preview[-1]["row"] == 20
