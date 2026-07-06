import openpyxl
from openpyxl.styles import Font

from app import comparator, workbook_writer
from app.workbook_reader import load_sheet

from .factories import (
    HEADER_ROW_1,
    HEADER_ROW_2,
    write_titled_two_row_workbook,
    write_two_row_workbook,
)


def test_complete_oms_row_preserved_with_blanks_and_negatives(tmp_path):
    current_path = tmp_path / "current.xlsx"
    oms_path = tmp_path / "oms.xlsx"
    write_two_row_workbook(current_path, [["HFILE", "H FILE 25MM", "PCS", None, 16, 24, 8]])
    write_two_row_workbook(oms_path, [["HFILE", "H FILE 25MM", "PCS", None, 16, 24, -8]])

    current = load_sheet(current_path, "Current File")
    oms = load_sheet(oms_path, "OMS File")
    result = comparator.compare(current, oms)

    output_path = tmp_path / "result.xlsx"
    workbook_writer.write_result(output_path, current, oms, result)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["DIFFERENCES"]
    assert [c.value for c in ws[1]] == HEADER_ROW_1
    assert [c.value for c in ws[2]] == HEADER_ROW_2
    assert [c.value for c in ws[3]] == ["HFILE", "H FILE 25MM", "PCS", None, 16, 24, -8]


def test_all_columns_and_order_preserved(tmp_path):
    current_path = tmp_path / "current.xlsx"
    oms_path = tmp_path / "oms.xlsx"
    write_two_row_workbook(current_path, [["A1", "Desc", "EA", 1, 1, 1, 1]])
    write_two_row_workbook(oms_path, [["A1", "Desc", "EA", 1, 1, 1, 9]])

    current = load_sheet(current_path, "Current File")
    oms = load_sheet(oms_path, "OMS File")
    result = comparator.compare(current, oms)

    output_path = tmp_path / "result.xlsx"
    workbook_writer.write_result(output_path, current, oms, result)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["DIFFERENCES"]
    assert ws.max_column == 7
    assert [c.value for c in ws[1]] == HEADER_ROW_1
    assert [c.value for c in ws[3]] == ["A1", "Desc", "EA", 1, 1, 1, 9]


def test_new_codes_sheet_created(tmp_path):
    current_path = tmp_path / "current.xlsx"
    oms_path = tmp_path / "oms.xlsx"
    write_two_row_workbook(current_path, [["A1", "Desc", "EA", 1, 1, 1, 1]])
    write_two_row_workbook(
        oms_path,
        [
            ["A1", "Desc", "EA", 1, 1, 1, 1],
            ["A2", "Desc2", "EA", 1, 1, 1, 1],
        ],
    )

    current = load_sheet(current_path, "Current File")
    oms = load_sheet(oms_path, "OMS File")
    result = comparator.compare(current, oms)

    output_path = tmp_path / "result.xlsx"
    workbook_writer.write_result(output_path, current, oms, result)

    wb = openpyxl.load_workbook(output_path)
    assert "NEW_CODES" in wb.sheetnames
    ws = wb["NEW_CODES"]
    assert [c.value for c in ws[3]] == ["A2", "Desc2", "EA", 1, 1, 1, 1]


def test_missing_from_oms_sheet_created_when_present(tmp_path):
    current_path = tmp_path / "current.xlsx"
    oms_path = tmp_path / "oms.xlsx"
    write_two_row_workbook(
        current_path,
        [
            ["A1", "Desc", "EA", 1, 1, 1, 1],
            ["A2", "Desc2", "EA", 1, 1, 1, 1],
        ],
    )
    write_two_row_workbook(oms_path, [["A1", "Desc", "EA", 1, 1, 1, 1]])

    current = load_sheet(current_path, "Current File")
    oms = load_sheet(oms_path, "OMS File")
    result = comparator.compare(current, oms)

    output_path = tmp_path / "result.xlsx"
    workbook_writer.write_result(output_path, current, oms, result)

    wb = openpyxl.load_workbook(output_path)
    assert "MISSING_FROM_OMS" in wb.sheetnames
    ws = wb["MISSING_FROM_OMS"]
    assert [c.value for c in ws[3]] == ["A2", "Desc2", "EA", 1, 1, 1, 1]


def test_missing_from_oms_sheet_absent_when_not_needed(tmp_path):
    current_path = tmp_path / "current.xlsx"
    oms_path = tmp_path / "oms.xlsx"
    write_two_row_workbook(current_path, [["A1", "Desc", "EA", 1, 1, 1, 1]])
    write_two_row_workbook(oms_path, [["A1", "Desc", "EA", 1, 1, 1, 9]])

    current = load_sheet(current_path, "Current File")
    oms = load_sheet(oms_path, "OMS File")
    result = comparator.compare(current, oms)

    output_path = tmp_path / "result.xlsx"
    workbook_writer.write_result(output_path, current, oms, result)

    wb = openpyxl.load_workbook(output_path)
    assert "MISSING_FROM_OMS" not in wb.sheetnames


def test_header_copied_correctly_when_source_has_report_title_block(tmp_path):
    current_path = tmp_path / "current.xlsx"
    oms_path = tmp_path / "oms.xlsx"
    write_titled_two_row_workbook(current_path, [["A1", "Desc", "EA", 1, 1, 1, 1]])
    write_titled_two_row_workbook(oms_path, [["A1", "Desc", "EA", 1, 1, 1, 9]])

    current = load_sheet(current_path, "Current File")
    oms = load_sheet(oms_path, "OMS File")
    result = comparator.compare(current, oms)

    output_path = tmp_path / "result.xlsx"
    workbook_writer.write_result(output_path, current, oms, result)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["DIFFERENCES"]
    assert [c.value for c in ws[1]] == HEADER_ROW_1
    assert [c.value for c in ws[2]] == HEADER_ROW_2
    assert [c.value for c in ws[3]] == ["A1", "Desc", "EA", 1, 1, 1, 9]


def test_styled_source_cells_can_be_saved(tmp_path):
    """Excel exports from real OMS reports have bold/colored cells, unlike
    the plain fixtures above - openpyxl wraps styled cells' font/fill/border
    in a StyleProxy that isn't hashable, so a naive style copy raises
    TypeError only when a source cell actually has a style applied."""
    current_path = tmp_path / "current.xlsx"
    oms_path = tmp_path / "oms.xlsx"

    for path, balance in [(current_path, 8), (oms_path, -8)]:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HEADER_ROW_1)
        ws.append(HEADER_ROW_2)
        ws.append(["HFILE", "H FILE 25MM", "PCS", None, 16, 24, balance])
        for cell in ws[3]:
            cell.font = Font(bold=True)
        wb.save(path)

    current = load_sheet(current_path, "Current File")
    oms = load_sheet(oms_path, "OMS File")
    result = comparator.compare(current, oms)

    output_path = tmp_path / "result.xlsx"
    workbook_writer.write_result(output_path, current, oms, result)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["DIFFERENCES"]
    assert [c.value for c in ws[3]] == ["HFILE", "H FILE 25MM", "PCS", None, 16, 24, -8]
    assert ws.cell(row=3, column=1).font.bold is True


def test_field_changes_sheet_lists_old_and_new_values(tmp_path):
    current_path = tmp_path / "current.xlsx"
    oms_path = tmp_path / "oms.xlsx"
    write_two_row_workbook(current_path, [["A1", "Desc", "EA", 1, 1, 1, 4]])
    write_two_row_workbook(oms_path, [["A1", "Desc", "EA", 1, 1, 1, 3]])

    current = load_sheet(current_path, "Current File")
    oms = load_sheet(oms_path, "OMS File")
    result = comparator.compare(current, oms)

    output_path = tmp_path / "result.xlsx"
    workbook_writer.write_result(output_path, current, oms, result)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["FIELD_CHANGES"]
    assert [c.value for c in ws[1]] == ["Code", "Column", "Old Value (Current File)", "New Value (OMS File)"]
    assert [c.value for c in ws[2]] == ["A1", "Balance", 4, 3]


def test_default_output_filename_format():
    name = workbook_writer.default_output_filename()
    assert name.startswith("Garg_Dental_Comparison_")
    assert name.endswith(".xlsx")
